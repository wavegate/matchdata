import json
import pandas
from pandas import DataFrame
import openpyxl
from fuzzywuzzy import fuzz, process

data = pandas.read_json('programs.json')

wb = openpyxl.load_workbook('compiled.xlsx')


count = 0
for ws in wb:
    if ws.title == "Surgery":
        names = []
        with open('programs.json') as f:
            data = json.load(f)
        for row in data:
            if row["specialty"] == "Surgery":
                names.append(row["program"][0])
        dictionary = {}
        for row in ws.iter_rows(min_row=2,min_col=2,max_col=2):
            for cell in row:
                #if count < 1:
                str2Match = cell.value
                if str2Match:
                    highest = process.extractOne(str2Match,names)
                    #highest = process.extractOne(str2Match, names, scorer=fuzz.partial_token_sort_ratio)
                    dictionary[str2Match] = highest[0]
                    print(str2Match + " : " + highest[0])
                    ws.cell(cell.row, cell.column+2).value = dictionary[str2Match]
    wb.save('names.xlsx')
           # count = count + 1
        
        #Ratios = process.extract(str2Match, names)
        #print(Ratios)
        #highest = process.extractOne(str2Match,names)
