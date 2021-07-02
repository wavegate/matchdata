import json
import pandas
from pandas import DataFrame
import openpyxl
from openpyxl import Workbook
from fuzzywuzzy import fuzz, process
from datetime import datetime

wb = openpyxl.load_workbook('surgery_dates.xlsx')
ws = wb.active
    
    
result_wb = Workbook()
result_ws = result_wb.active

def date_to_string(date):
    return date.strftime("%m/%d/%Y")

for row in ws.iter_rows(min_row=5):
    rindex = row[0].row
    result_ws.cell(rindex-4,1).value = row[0].value
    list_of_offer_dates = []
    for cell in row[7:11]:
        if isinstance(cell.value, datetime):
            if cell.value.month < 6:
                cell.value = cell.value.replace(year=2021)
            list_of_offer_dates.append(cell.value)
    list_of_interview_dates = []
    for cell in row[13:33]:
        if isinstance(cell.value, datetime):
            if cell.value.month < 6:
                cell.value = cell.value.replace(year=2021)
            list_of_interview_dates.append(cell.value)
    result_ws.cell(rindex-4,2).value = ",".join(map(date_to_string, list_of_offer_dates))
    result_ws.cell(rindex-4,3).value = ",".join(map(date_to_string, list_of_interview_dates))

result_wb.save('surgery_dates_output.xlsx')
