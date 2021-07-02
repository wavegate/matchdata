import json
import pandas
from pandas import DataFrame
import openpyxl
from fuzzywuzzy import fuzz, process

wb = openpyxl.load_workbook('names.xlsx')

count = 0
for ws in wb:
    if ws.title == "Surgery":
        for col in ws.iter_cols(min_row=3,min_col=4,max_col=4):
            for cell in col:
                if ws.cell(cell.row,cell.column+1).value == "X":
                    ws.cell(cell.row,cell.column+3).value = ws.cell(cell.row,cell.column+2).value
                else:
                    ws.cell(cell.row,cell.column+3).value = cell.value
    wb.save('names_compiled.xlsx')
           # count = count + 1
        
        #Ratios = process.extract(str2Match, names)
        #print(Ratios)
        #highest = process.extractOne(str2Match,names)

