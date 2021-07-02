import json
import openpyxl
from fuzzywuzzy import fuzz, process

wb = openpyxl.load_workbook('gensurg2021.xlsx')

names = []
        with open('./programs.json') as f:
            data = json.load(f)
        for row in data:
            if row["specialty"] == "Surgery":
                names.append(row["program"][0])
                
for ws in wb:
    if ws.title == "2020 Interview Impressions":
        for col in ws.iter_cols(min_row=4,min_col=1,max_col=1):
            for cell in col:
                if ws.cell(cell.row,cell.column+1).value == "X":
                    ws.cell(cell.row,cell.column+3).value = ws.cell(cell.row,cell.column+2).value
                else:
                    ws.cell(cell.row,cell.column+3).value = cell.value

wb.save('names_compiled.xlsx')
           # count = count + 1
        
        #Ratios = process.extract(str2Match, names)
        #print(Ratios)
        #highest = process.extractOne(str2Match,names)


