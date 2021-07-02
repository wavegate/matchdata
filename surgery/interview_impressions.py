import json
import openpyxl
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook

wb = openpyxl.load_workbook('interview_impressions2.xlsx')

names = []
with open('programs.json') as f:
    data = json.load(f)
for row in data:
    if row["specialty"] == "Surgery":
        names.append(row["program"][0])
        
result_wb = openpyxl.load_workbook('interview_impressions_processed2.xlsx')
result_ws = result_wb.active
                
ws = wb.active
for row in ws.iter_rows(min_row=4,max_row=313):
    rindex = row[0].row
    highest = []
    name = None
    if row[0].value:
        print(str(rindex-3) + " : " + row[0].value)
        highest = process.extract(row[0].value,names)
        if highest[0][1] > 89:
            name = highest[0][0]
    #if not result_ws.cell(rindex-3,1).value:
    #    else:
    #        highest = process.extract(row[0].value,names,scorer=fuzz.partial_ratio)
    #        print(highest)
    #        select = input()
    #        if select.isdigit():
    #            name = highest[int(select)-1][0]
    #            print(name)
    #        else:
    #            name = select
        result_ws.cell(rindex-3,1).value=name
    list_of_reviews = []
    for cell in row[5:8]:
        if cell.value:
            list_of_reviews.append(cell.value)
    result_ws.cell(rindex-3,2).value = "XXXXX".join(list_of_reviews)
    result_wb.save('interview_impressions_processed2.xlsx')
    print("\n")



